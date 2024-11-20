# Ian Bullard
# UWYO COSC 1010
# 11/19/24
# HW 05
# Lab Section: 14
# Sources, people worked with, help given to: Bounced ideas around with Patrick McGrath
# Asked ChatGPT how to get the AA-AF columns to work becasuse that was the only stuff not printing correctly
import openpyxl

from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter
import string
wb = openpyxl.Workbook()
sheet = wb.active

blue = Color(rgb ='90d5ff')
fill_bl = PatternFill(patternType='solid',fgColor = blue)

dark_brown =  Color(rgb ="492f24")
fill_db = PatternFill(patternType='solid',fgColor = dark_brown)

light_brown = Color(rgb = "895129")
fill_lb = PatternFill(patternType='solid',fgColor = light_brown)

yellow = Color(rgb = 'ffde21')
fill_y = PatternFill(patternType='solid',fgColor = yellow)

light_grey = Color(rgb = 'd3d3d3')
fill_lg = PatternFill(patternType='solid',fgColor = light_grey)

cells_db = ['A36','A37','A38','B34','B35','B36','C31','C32','C33','C34','C35','D29','D30','D31','E27','E28','E29','E30','F25','F26','F27','G23','G24','G25','G26','H24','H25','P37','P38','Q35','Q36','Q37','Q38','R33','R34','R35','S30','S31','S32','S33','S34','T29','T30','T31','U27','U28','U29','U30','V24','V25','V26','V27']
cells_lb = ['B37','B38','C37','C38','D37','D38','E37','E38','F37','F38','G37','G38','H37','H38','I37','I38','J37','J38','K37','K38','L37','L38','M37','M38','N37','N38','O37','O38','R37','R38','S37','S38','T37','T38','U37','U38','V37','V38','W37','W38','X37','X38','Y37','Y38','Z37','Z38','AA37','AA38','AB37','AB38','AC37','AC38','AD37','AD38','AE37','AE38','AF38','C36','D36','E36','F36','G36','H36','I36','J36','K36','L36','M36','N36','O36','P36','R36','S36','T36','U36','V36','W36','X36','Y36','Z36','AA36','AB36','AC36','AD36','AE36','D35','E35','F35','G35','H35','I35','J35','K35','L35','M35','N35','O35','S35','T35','U35','V35','W35','X35','Y35','Z35','AA35','AB35','AC35','AD35','D34','E34','F34','G34','H34','I34','J34','K34','L34','M34','N34','O34','T34','U34','V34','W34','X34','Y34','Z34','AA34','AB34','AC34','AD34','D33','E33','F33','G33','H33','I33','J33','K33','L33','M33','N33','T33','U33','V33','W33','X33','Y33','Z33','AA33','AB33','AC33','D32','E32','F32','G32','H32','I32','J32','K32','L32','M32','N32','T32','U32','V32','W32','X32','Y32','Z32','AA32','AB32','AC32','E31','F31','G31','H31','I31','J31','K31','L31','M31','U31','V31','W31','X31','Y31','Z31','AA31','AB31','F30','G30','H30','I30','J30','K30','L30','M30','V30','W30','X30','Y30','Z30','AA30','F29','G29','H29','I29','J29','K29','L29','M29','V29','W29','X29','Y29','Z29','AA29','F28','G28','H28','I28','J28','K28','L28','V28','W28','X28','Y28','Z28','AA28','G27','H27','I27','J27','K27','L27','W27','X27','Y27','Z27','H26','I26','J26','K26','W26','X26','Y26','W25','X25','Y25','X24']
cells_y = ['Y5','Y6','Z4','Z5','Z6','Z7','AA3','AA4','AA5','AA6','AA7','AA8','AB2','AB3','AB4','AB5','AB6','AB7','AB8','AB9','AC3','AC4','AC5','AC6','AC7','AC8','AD4','AD5','AD6','AD7','AE5','AE6']
cells_lg = ['E7','F6','F7','F8','G5','G6','G7','G8','G9','H5','H6','H7','H8','H9','I6','I7','I8','J7','P11','Q10','Q11','Q12','R9','R10','R11','R12','R13','U11','T10','T11','T12','S9','S10','S11','S12','S13','H21','H22','H23','I20','I21','I22','I23','I24','I25','J21','J22','J23','J24','J25','K23','K24','K25','V23','W22','W23','W24','X21','X22','X23','Y22','Y23','Y24','Z24','Z25','Z26']

columns = [chr(i) for i in range(65, 91)]
columns += {f"A{chr}" for chr in 'ABCDEF'}

for chr in columns:
    sheet.column_dimensions[chr].width = 5
    for i in range(1,39):
        sheet.row_dimensions[i].height= 20
        coord = chr+str(i)
        if coord in cells_db:
            sheet[coord].fill = fill_db
        elif coord in cells_lb:
            sheet[coord].fill = fill_lb
        elif coord in cells_y:
            sheet[coord].fill = fill_y
        elif coord in cells_lg:
            sheet[coord].fill = fill_lg
        else:
            sheet[coord].fill = fill_bl
# for chr in string.ascii_uppercase[:33]:
#     sheet.column_dimensions[chr].width = 5
#     for i in range(1,39):
#         sheet.row_dimensions[i].height= 20
#         coord = chr+str(i)
#         if coord in cells_db:
#             sheet[coord].fill = fill_db
#         elif coord in cells_lb:
#             sheet[coord].fill = fill_lb
#         elif coord in cells_y:
#             sheet[coord].fill = fill_y
#         elif coord in cells_lg:
#             sheet[coord].fill = fill_lg
#         else:
#             sheet[coord].fill = fill_bl

wb.save("colors.xlsx")